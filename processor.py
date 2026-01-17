import pandas as pd
import re
from openpyxl.styles import Alignment, PatternFill, Border, Side

def extract_area_logic(text):
    if pd.isna(text) or text == "": return 0.0
    
    # Cleanup: normalize spaces and remove commas in numbers (e.g., 13,600 -> 13600)
    text = " ".join(str(text).split())
    text = re.sub(r'(\d),(\d)', r'\1\2', text)
    text = text.replace(' ,', ',').replace(', ', ',')
    
    # Focus Logic: Jump past land survey details to unit details
    focus_keywords = r'(?:इमारतीमधील|अपार्टमेंटमधील|सदनिका|फ्लॅट|युनिट|टावर|टॉवर|flat|unit|tower)'
    parts = re.split(focus_keywords, text, flags=re.IGNORECASE)
    relevant_text = " ".join(parts[1:]) if len(parts) > 1 else text

    # Regex for Units
    m_unit = r'(?:चौ\.?\s*मी\.?|चौरस\s*मी[टत]र|sq\.?\s*m(?:tr)?\.?)'
    f_unit = r'(?:चौ\.?\s*फू\.?|चौरस\s*फु[टत]|sq\.?\s*f(?:t)?\.?)'
    total_keywords = r'(?:ए[ककु]ण\s*क्षेत्र|क्षेत्रफळ|total\s*area|सेलेबल\s*क्षेत्र|एकूण\s*सेलेबल)'
    parking_keywords = ["पार्किंग", "पार्कींग", "parking", "पार्कीग", "पार्किंगसह", "कार पार्क"]

    # METRIC SUMMATION (Collecting all components for duplexes/luxury units)
    m_segments = re.split(f'(\d+\.?\d*)\s*{m_unit}', relevant_text, flags=re.IGNORECASE)
    m_vals = []
    for i in range(1, len(m_segments), 2):
        val = float(m_segments[i])
        context_before = m_segments[i-1].lower()
        if 0 < val < 1200 and not any(word in context_before for word in parking_keywords):
            m_vals.append(val)
    
    if m_vals:
        # Avoid double-counting if the final number is the sum of previous ones
        if len(m_vals) > 1 and abs(m_vals[-1] - sum(m_vals[:-1])) < 1.0:
            return round(m_vals[-1], 3)
        return round(sum(m_vals), 3)

    # IMPERIAL FALLBACK
    f_segments = re.split(f'(\d+\.?\d*)\s*{f_unit}', relevant_text, flags=re.IGNORECASE)
    f_vals = []
    for i in range(1, len(f_segments), 2):
        val = float(f_segments[i])
        context_before = f_segments[i-1].lower()
        if 0 < val < 15000 and not any(word in context_before for word in parking_keywords):
            f_vals.append(val)
                
    if f_vals:
        if len(f_vals) > 1 and abs(f_vals[-1] - sum(f_vals[:-1])) < 10:
            return round(f_vals[-1] / 10.764, 3)
        return round(sum(f_vals) / 10.764, 3)
        
    return 0.0

def determine_config(area_sqft, t1, t2, t3):
    if area_sqft == 0: return "N/A"
    if area_sqft < t1: return "1 BHK"
    elif area_sqft < t2: return "2 BHK"
    elif area_sqft < t3: return "3 BHK"
    else: return "4 BHK+"

def apply_excel_formatting(df, writer, sheet_name, is_summary=True):
    df.to_excel(writer, sheet_name=sheet_name, index=False)
    worksheet = writer.sheets[sheet_name]
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    for i in range(1, worksheet.max_row + 1):
        for j in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=i, column=j)
            cell.alignment = center_align
            if is_summary: cell.border = thin_border
