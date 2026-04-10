import streamlit as st
import pandas as pd
import re
import io
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email.utils import formataddr
from email import encoders
from openpyxl.styles import Alignment, PatternFill, Border, Side
from groq import Groq
import os
import json
import time

# --- EMAIL CONFIGURATION ---
SENDER_EMAIL = "atharvaujoshi@gmail.com"
SENDER_NAME = "Spydarr Market Research"
APP_PASSWORD = "nybl zsnx zvdw edqr"

# llama3-8b-8192 = 500k tokens/day free tier (vs 100k for 70b)
GROQ_MODEL = "llama3-8b-8192"

# --- GROQ CLIENT ---
@st.cache_resource
def get_groq_client():
    api_key = os.environ.get("GROQ_API_KEY") or st.secrets.get("GROQ_API_KEY", None)
    if not api_key:
        st.error("GROQ_API_KEY not found. Add it to your .env or Streamlit secrets.")
        st.stop()
    return Groq(api_key=api_key)


# ─────────────────────────────────────────────────────────────
# STEP 1 — Pre-filter: extract only the flat-relevant portion
# This runs locally (no API call) and cuts tokens by ~80%
# ─────────────────────────────────────────────────────────────
def prefilter_description(text: str) -> str:
    """
    From a full property description, extract only the sentence/clause
    that describes the flat's areas (carpet + balcony + utility + terrace).
    Removes land survey lines, RERA numbers, stamp duty notes, etc.
    Returns a short cleaned string to send to Groq.
    """
    if pd.isna(text) or str(text).strip() == "":
        return ""

    text = str(text)

    # Split on common clause separators
    # We want clauses that contain area keywords for the flat
    flat_keywords = [
        "कारपेट", "कार्पेट", "carpet",
        "बाल्कनी", "बालकनी", "balcony",
        "युटिलिटी", "utility",
        "टेरेस", "terrace",
        "ड्राय", "dry",
        "फ्लॅट नं", "सदनिका नं", "अपार्टमेंट नं",
        "flat no", "apt no",
    ]

    # Split on sentence-like boundaries
    parts = re.split(r'[,।\n]|\.\s+', text)

    kept = []
    for part in parts:
        part = part.strip()
        if not part:
            continue
        # Skip pure land/survey lines
        if re.search(r'स\.नं\.|सर्व्हे नं|हिस्सा नं|हे \d+|आर,|यापैकी क्षेत्र \d{4,}|एकूण क्षेत्र \d{4,}', part):
            continue
        # Skip RERA / stamp duty / legal boilerplate
        if re.search(r'रेरा|RERA|मुद्रांक|शासन आदेश|Survey Number|महिला खरेदी|सवलत', part, re.IGNORECASE):
            continue
        # Keep if it has flat area keywords
        if any(kw.lower() in part.lower() for kw in flat_keywords):
            kept.append(part)
        # Also keep the flat identifier line (floor, flat number)
        elif re.search(r'मजल्यावरील|फ्लॅट नं|सदनिका नं|अपार्टमेंट नं', part):
            kept.append(part)

    result = ", ".join(kept) if kept else text[:600]
    return result[:1200]  # hard cap — never send more than 1200 chars


SYSTEM_PROMPT = """You are an expert at reading Indian property registration documents in Marathi and English.

Extract ONLY the flat/apartment area components and return their SUM in sq.mt.

INCLUDE:
- Carpet area: कारपेट, कार्पेट, carpet area
- Any balcony: बाल्कनी, बालकनी, ओपन बाल्कनी, अटॅच बाल्कनी, एन्क्लोज बाल्कनी, लगतेच बाल्कनी
- Dry balcony: ड्राय बाल्कनी, ड्राय बालकनी
- Utility: युटिलिटी, युटिलिटी बालकनी
- Flat terrace: टेरेस (only if listed beside carpet/balcony, NOT open-to-sky)

EXCLUDE:
- Parking: पार्किंग, कार पार्किंग, पार्कींग, कव्हर्ड कार पार्किंग (has stall number)
- Open terrace / sky: ओपन टेरेस, ओपन टू स्काय
- Land areas: anything with हे/आर units, or survey number land extents

UNITS:
- चौ.मी. / sq.mt → use directly
- चौ.फूट / चौ.फुट / चौ.फु / sq.ft → divide by 10.764
- If both units given for same item, use चौ.मी. value only

Return ONLY raw JSON, no markdown, no explanation:
{"components":[{"label":"carpet","value_sqmt":64.61},{"label":"balcony","value_sqmt":5.99}],"total_sqmt":70.60}
If nothing found: {"components":[],"total_sqmt":0.0}"""


def call_groq_once(client, text):
    """Single Groq API call. Returns (raw, parsed, error)."""
    try:
        response = client.chat.completions.create(
            model=GROQ_MODEL,
            messages=[
                {"role": "system", "content": SYSTEM_PROMPT},
                {"role": "user", "content": text}
            ],
            temperature=0,
            max_tokens=256,
        )
        raw = response.choices[0].message.content.strip()
        clean = re.sub(r"```json|```", "", raw).strip()
        m = re.search(r'\{.*\}', clean, re.DOTALL)
        if m:
            clean = m.group(0)
        parsed = json.loads(clean)
        return raw, parsed, None
    except json.JSONDecodeError as e:
        return locals().get('raw', ''), None, f"JSON error: {e}"
    except Exception as e:
        return "", None, f"API error: {e}"


def extract_area_groq(text, client, debug_log=None):
    """Pre-filter description, then call Groq. Retry up to 3x on rate limit."""
    if pd.isna(text) or str(text).strip() == "":
        return 0.0

    filtered = prefilter_description(text)

    log_entry = {
        "original_length": len(str(text)),
        "filtered_length": len(filtered),
        "filtered_text": filtered,
        "raw_response": "",
        "components": [],
        "total_sqmt": 0.0,
        "status": "ok"
    }

    result_total = 0.0

    for attempt in range(3):
        raw, parsed, error = call_groq_once(client, filtered)
        log_entry["raw_response"] = raw

        if error is None and parsed is not None:
            result_total = round(float(parsed.get("total_sqmt", 0.0)), 3)
            log_entry["components"] = parsed.get("components", [])
            log_entry["total_sqmt"] = result_total
            log_entry["status"] = "ok"
            break
        elif error and "429" in str(error):
            # Rate limited — extract wait time and sleep
            wait_match = re.search(r'try again in (\d+)m(\d+)', str(error))
            if wait_match:
                wait_secs = int(wait_match.group(1)) * 60 + int(wait_match.group(2)) + 5
            else:
                wait_secs = 30
            log_entry["status"] = f"rate limited, waiting {wait_secs}s (attempt {attempt+1})"
            if debug_log is not None:
                debug_log.append(dict(log_entry))
            time.sleep(min(wait_secs, 120))  # cap at 2 min
        else:
            log_entry["status"] = f"attempt {attempt+1} failed: {error}"
            time.sleep(2)

        if attempt == 2:
            # Fallback: grab last plausible float from raw
            nums = re.findall(r'\b\d{1,3}\.\d{1,3}\b', raw)
            plausible = [float(n) for n in nums if 2.0 < float(n) < 500.0]
            result_total = round(plausible[-1], 3) if plausible else 0.0
            log_entry["total_sqmt"] = result_total
            log_entry["status"] += f" | fallback={result_total}"

    if debug_log is not None:
        debug_log.append(log_entry)

    return result_total


def send_email(recipient_email, excel_data, filename):
    try:
        recipient_name = recipient_email.split('@')[0].replace('.', ' ').title()
        msg = MIMEMultipart()
        msg['From'] = formataddr((SENDER_NAME, SENDER_EMAIL))
        msg['To'] = recipient_email
        msg['Subject'] = "Spydarr Market Research Summary"
        body = f"""Dear {recipient_name},

Please find the attached professional property analysis report generated by the dashboard.

The report includes:
1. Raw Data with calculated APR and Configurations.
2. A summarized view of APR statistics across properties.

Regards,
Atharva Joshi"""
        msg.attach(MIMEText(body, 'plain'))
        part = MIMEBase('application', 'octet-stream')
        part.set_payload(excel_data)
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', f"attachment; filename={filename}")
        msg.attach(part)
        server = smtplib.SMTP('smtp.gmail.com', 587)
        server.starttls()
        server.login(SENDER_EMAIL, APP_PASSWORD)
        server.send_message(msg)
        server.quit()
        return True
    except Exception as e:
        st.error(f"Error sending email: {e}")
        return False


def determine_config(area, t1, t2, t3):
    if area == 0: return "N/A"
    if area < t1: return "1 BHK"
    elif area < t2: return "2 BHK"
    elif area < t3: return "3 BHK"
    else: return "4 BHK"


def apply_excel_formatting(df, writer, sheet_name, is_summary=True):
    df.to_excel(writer, sheet_name=sheet_name, index=False)
    worksheet = writer.sheets[sheet_name]
    worksheet.freeze_panes = "A2"
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    colors = ["A2D2FF", "FFD6A5", "CAFFBF", "FDFFB6", "FFADAD", "BDB2FF", "9BF6FF"]

    for i in range(1, worksheet.max_row + 1):
        for j in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=i, column=j)
            cell.alignment = center_align
            if is_summary:
                cell.border = thin_border

    if is_summary:
        color_idx, start_row_prop = 0, 2
        start_row_loc = 2
        last_col = len(df.columns)
        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

        for i in range(2, len(df) + 3):
            curr_loc = df.iloc[i-2, 0] if i-2 < len(df) else None
            prev_loc = df.iloc[i-3, 0] if i-3 >= 0 else None
            curr_prop = df.iloc[i-2, 1] if i-2 < len(df) else None
            prev_prop = df.iloc[i-3, 1] if i-3 >= 0 else None

            if curr_prop != prev_prop and i > 2:
                fill = PatternFill(start_color=colors[color_idx % len(colors)],
                                   end_color=colors[color_idx % len(colors)], fill_type="solid")
                for r in range(start_row_prop, i):
                    for c in range(2, last_col + 1):
                        worksheet.cell(row=r, column=c).fill = fill
                if i-1 > start_row_prop:
                    worksheet.merge_cells(start_row=start_row_prop, start_column=2,
                                          end_row=i-1, end_column=2)
                    worksheet.merge_cells(start_row=start_row_prop, start_column=last_col,
                                          end_row=i-1, end_column=last_col)
                start_row_prop = i
                color_idx += 1

            if curr_loc != prev_loc and i > 2:
                for r in range(start_row_loc, i):
                    worksheet.cell(row=r, column=1).fill = white_fill
                if i-1 > start_row_loc:
                    worksheet.merge_cells(start_row=start_row_loc, start_column=1,
                                          end_row=i-1, end_column=1)
                start_row_loc = i


# ─────────────────────────────────────────
# STREAMLIT UI
# ─────────────────────────────────────────
st.set_page_config(page_title="Spydarr Dashboard", layout="wide")
st.title("Spydarr Dashboard")
st.markdown(
    "<div style='margin-top:-15px;margin-bottom:5px;'>"
    "<span style='background-color:#FFFF00;padding:2px 8px;border-radius:4px;"
    "border:1px solid #E6E600;font-size:0.9em;color:black;'>"
    "<u><strong>NOTE :-</strong> Please cross-check the report manually.</u></span></div>",
    unsafe_allow_html=True
)
st.markdown("[Property Report Tool · Streamlit](https://summarybeyondwalls.streamlit.app/)")
st.divider()

# ── Sidebar ──────────────────────────────
st.sidebar.header("Calculation Settings")
loading_factor = st.sidebar.number_input("Loading Factor", min_value=1.0, value=1.35, step=0.001, format="%.3f")
t1 = st.sidebar.number_input("1 BHK Threshold (sq.ft)", value=600)
t2 = st.sidebar.number_input("2 BHK Threshold (sq.ft)", value=850)
t3 = st.sidebar.number_input("3 BHK Threshold (sq.ft)", value=1100)

st.sidebar.divider()
st.sidebar.header("🔧 Groq API Test")
st.sidebar.caption("Run this FIRST to confirm your API key and model work.")
if st.sidebar.button("▶ Test Groq API"):
    test_text = 'फ्लॅट नं. 402, कारपेट क्षेत्र 64.61 चौ. मी., ओपन बाल्कनी क्षेत्र 5.99 चौ.मी., कव्हर्ड कार पार्किंग नं.(जी एल - 60), क्षेत्र 12.50 चौ. मी.'
    try:
        test_client = get_groq_client()
        raw, parsed, error = call_groq_once(test_client, test_text)
        if error:
            st.sidebar.error(f"❌ {error}")
        else:
            total = parsed.get('total_sqmt')
            ok = abs(total - 70.60) < 1.0
            if ok:
                st.sidebar.success(f"✅ API OK! Got {total} sq.mt (expected ~70.60)")
            else:
                st.sidebar.warning(f"⚠️ API responded but got {total} sq.mt (expected ~70.60) — check prompt")
            st.sidebar.code(raw, language='json')
    except Exception as e:
        st.sidebar.error(f"❌ {e}")

# ── File Upload ───────────────────────────
uploaded_file = st.file_uploader("Upload Data File", type=["xlsx", "csv"])

if uploaded_file:
    df = pd.read_csv(uploaded_file) if uploaded_file.name.endswith('.csv') else pd.read_excel(uploaded_file)
    clean_cols = {c.lower().strip(): c for c in df.columns}
    desc_col = clean_cols.get('property description')
    cons_col = clean_cols.get('consideration value')
    prop_col = clean_cols.get('property')
    date_col = clean_cols.get('completion date')
    loc_col  = clean_cols.get('micromarket')

    if desc_col and cons_col and prop_col and date_col and loc_col:
        client = get_groq_client()

        # Estimate token usage upfront
        total_rows = len(df)
        avg_filtered_len = 400  # chars after prefilter, ~100 tokens each
        est_tokens = total_rows * (avg_filtered_len // 4 + 300)  # +300 for system prompt
        st.info(
            f"📊 **{total_rows} rows** detected. "
            f"Estimated token usage: ~**{est_tokens:,}** tokens "
            f"(model limit: 500k/day on free tier). "
            f"{'⚠️ May hit daily limit — consider splitting file.' if est_tokens > 400000 else '✅ Should fit within free tier.'}"
        )

        with st.spinner('Extracting carpet areas via Groq LLM...'):
            areas = []
            debug_log = []
            progress = st.progress(0, text="Starting...")
            token_counter = st.empty()

            for idx, row in df.iterrows():
                area = extract_area_groq(row[desc_col], client, debug_log=debug_log)
                areas.append(area)
                pct = (idx + 1) / total_rows
                zeros_so_far = sum(1 for a in areas if a == 0.0)
                progress.progress(pct, text=f"Row {idx+1}/{total_rows} — last: {area} sq.mt | zeros so far: {zeros_so_far}")

            progress.empty()
            token_counter.empty()
            df['Carpet Area (SQ.MT)'] = areas

        with st.spinner('Calculating APR and generating report...'):
            df['Carpet Area (SQ.FT)'] = (df['Carpet Area (SQ.MT)'] * 10.764).round(3)
            df['Saleable Area'] = (df['Carpet Area (SQ.FT)'] * loading_factor).round(3)
            df['APR'] = df.apply(
                lambda r: round(r[cons_col] / r['Saleable Area'], 3) if r['Saleable Area'] > 0 else 0,
                axis=1
            )
            df['Configuration'] = df['Carpet Area (SQ.FT)'].apply(
                lambda x: determine_config(x, t1, t2, t3)
            )
            df[date_col] = pd.to_datetime(df[date_col], errors='coerce')

            valid_df = df[df['Carpet Area (SQ.FT)'] > 0].sort_values(
                [loc_col, prop_col, 'Configuration', 'Carpet Area (SQ.FT)']
            )

            project_counts = valid_df.groupby(prop_col).size().reset_index(name='Total Count')
            summary = valid_df.groupby(
                [loc_col, prop_col, 'Configuration', 'Carpet Area (SQ.FT)']
            ).agg(
                Last_Date=(date_col, 'max'),
                Min_APR=('APR', 'min'),
                Max_APR=('APR', 'max'),
                Avg_APR=('APR', 'mean'),
                Median_APR=('APR', 'median'),
                Property_Count=(prop_col, 'count')
            ).reset_index()

            summary = summary.merge(project_counts, on=prop_col, how='left')
            summary['Last_Date'] = pd.to_datetime(summary['Last_Date'], errors='coerce')
            summary['Last_Date'] = summary['Last_Date'].apply(
                lambda x: x.strftime('%b-%Y') if pd.notnull(x) else "N/A"
            )
            summary.columns = [
                'Location', 'Property', 'Configuration', 'Carpet Area(SQ.FT)',
                'Last Completion Date', 'Min. APR', 'Max APR',
                'Average of APR', 'Median of APR', 'Count of Property', 'Total Count'
            ]
            summary = summary[[
                'Location', 'Property', 'Last Completion Date', 'Configuration',
                'Carpet Area(SQ.FT)', 'Min. APR', 'Max APR',
                'Average of APR', 'Median of APR', 'Count of Property', 'Total Count'
            ]]

            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                apply_excel_formatting(df, writer, 'Raw Data', is_summary=False)
                apply_excel_formatting(summary, writer, 'Summary', is_summary=True)

        zero_count = sum(1 for a in areas if a == 0.0)
        st.success(f"✅ Done! {total_rows - zero_count}/{total_rows} rows extracted successfully.")
        if zero_count > 0:
            st.warning(f"⚠️ {zero_count} rows returned 0 — open Debug expander to investigate.")

        # Preview table
        with st.expander("🔍 Preview: Carpet Area Extraction (first 20 rows)"):
            preview = df[[desc_col, 'Carpet Area (SQ.MT)', 'Carpet Area (SQ.FT)']].head(20).copy()
            preview['Carpet Area (SQ.MT)'] = preview['Carpet Area (SQ.MT)'].apply(
                lambda x: f"🔴 {x}" if x == 0.0 else f"🟢 {x}"
            )
            st.dataframe(preview, use_container_width=True)

        # Debug expander — zeros first
        with st.expander(f"🐛 Debug: LLM Responses ({zero_count} zeros)"):
            sorted_log = sorted(debug_log, key=lambda x: x['total_sqmt'])
            for entry in sorted_log:
                icon = "🔴" if entry['total_sqmt'] == 0.0 else "🟢"
                st.markdown(
                    f"{icon} **{entry['total_sqmt']} sq.mt** | "
                    f"Original: {entry['original_length']} chars → Filtered: {entry['filtered_length']} chars | "
                    f"Status: `{entry['status']}`"
                )
                with st.container():
                    col1, col2 = st.columns(2)
                    with col1:
                        st.caption("📤 Sent to Groq:")
                        st.code(entry['filtered_text'], language='text')
                    with col2:
                        st.caption("📥 Groq Response:")
                        st.code(entry['raw_response'] or "(empty)", language='json')
                st.divider()

        # Email + Download
        recipient = st.text_input("Recipient Name", placeholder="firstname.lastname")
        if st.button("Send to Email") and recipient:
            full_email = f"{recipient.strip().lower()}@beyondwalls.com"
            if send_email(full_email, output.getvalue(), "Spydarr_Market_Report.xlsx"):
                st.success(f"✅ Report sent to {full_email}")

        st.download_button(
            label="⬇️ Download Report",
            data=output.getvalue(),
            file_name="Spydarr_Market_Report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    else:
        st.error(
            "Missing required columns. Ensure file has: "
            "'Micromarket', 'Property Description', 'Consideration Value', 'Property', 'Completion Date'."
        )
